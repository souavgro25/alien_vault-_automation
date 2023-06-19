import openpyxl
from server_mails import send_email 
# from server_cats import login_cats


def get_department_name(file_path, ip_address):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find the column index of the IP address and department name columns
    ip_column_index = None
    department_column_index = None
    for column in sheet.iter_cols():
        for cell in column:
            if cell.value == "IPAddress":
                ip_column_index = cell.column
            elif cell.value == "Department":
                department_column_index = cell.column

    # If the IP address or department column is not found, return None
    if ip_column_index is None or department_column_index is None:
        return None

    # Iterate through the rows and search for the IP address
    for row in sheet.iter_rows(min_row=2):
        ip_cell = row[ip_column_index - 1]
        if ip_cell.value == ip_address:
            department_cell = row[department_column_index - 1]
            return department_cell.value

    octets = ip_address.split('.')
    if len(octets) >= 3:
        partial_ip = '.'.join(octets[:3])
        for row in sheet.iter_rows(min_row=2):
            ip_cell = row[ip_column_index - 1]
            if ip_cell.value and ip_cell.value.startswith(partial_ip):
                department_cell = row[department_column_index - 1]
                return department_cell.value

    # If the IP address is not found, return None

    return None


def find_department_email_recipients(department_name, department_file_path):
    # Load the department file
    department_workbook = openpyxl.load_workbook(department_file_path)
    department_sheet = department_workbook.active

    recipients = []

    # Find the department name in the department file
    for row in department_sheet.iter_rows(values_only=True):
        if department_name in row:
            department_column_index = row.index(department_name)

            # Retrieve all the row values of the department column
            for row_value in department_sheet.iter_rows(min_row=2, min_col=department_column_index + 1,
                                                        max_col=department_column_index + 1,
                                                        values_only=True):
                for recipient in row_value:
                    if recipient is not None:
                        recipients.append(recipient)
                    else:
                        break
            break

    return recipients


# Example usage
ip_file_path = "eGov.xlsx"
ip_address = "10.1.51.47"
department_file_path = "Department_list.xlsx"

# if department_name is not None:
#     print(f"The department name for IP address {ip_address} is {department_name}.")
# else:
#     print(f"No department name found for IP address {ip_address}.")

# for recipient in recipients:
#     print(recipient)


def create_host_ip_dictionary(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the sheet by name
    sheet = workbook[sheet_name]

    # Create an empty dictionary
    host_ip_dict = {}

    # Iterate over all rows until all columns are empty
    for row in sheet.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            break

        # Get the values in the row as a list
        values = list(row)

        # Extract the hostname and IP address from the values list
        host = values[0]
        ip = values[1] if len(values) > 1 else None

        # Add the hostname and IP address to the dictionary
        if host is not None:
            host_ip_dict[host] = ip

    return host_ip_dict


# Example usage
file_path = "ticketsopen.xlsx"
sheet_name = "Sheet1"

host_ip_dictionary = create_host_ip_dictionary(file_path, sheet_name)
print("Host-IP Dictionary:")
for host, ip in host_ip_dictionary.items():
    department_name = get_department_name(ip_file_path, ip)
    recipients = find_department_email_recipients(
        department_name, department_file_path)
    print(f"Hostname: {host}, IP: {ip} , department name : {department_name}")
    cc = ["egov@in2ittech.co.za"]
    account_name = "egov@in2ittech.co.za"  # Specify the desired email account name
    subject = f"Server Disconnected - Server Name: {host} and IP: {ip}"
    body = f"""
Hi Good day Team,

We have observed that Server Name: {host} and IP: {ip} which belongs to 
Department: {department_name} is disconnected on SIEM console.

The communication between the server and the SIEM (OSSIEM) is not working.Please check the 
SIEM (OSSIEM) agent on the Task bar and also confirm if SIEM (OSSIEM) services are running.

Please check and provide us an update for the same.\n

Thanks & Regards,
Sourabh Grover
EGov SOC team
0110546980
egov@in2ittech.co.za
    """
    # login_cats(subject,body,department_name)

    send_email(subject, body, recipients, cc, account_name)    
