import openpyxl
# from server_mails import send_email 
import sys


def get_department_name(file_path, ip_address):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find the column index of the IP address and department name columns
    ip_column_index = None
    department_column_index = None
    for column in sheet.iter_cols():
        for cell in column:
            if cell.value == "IPAddress":
                ip_column_index = cell.column
            elif cell.value == "Department":
                department_column_index = cell.column

    # If the IP address or department column is not found, return None
    if ip_column_index is None or department_column_index is None:
        return None

    # Iterate through the rows and search for the IP address
    for row in sheet.iter_rows(min_row=2):
        ip_cell = row[ip_column_index - 1]
        if ip_cell.value == ip_address:
            department_cell = row[department_column_index - 1]
            return department_cell.value

    octets = ip_address.split('.')
    if len(octets) >= 3:
        partial_ip = '.'.join(octets[:3])
        for row in sheet.iter_rows(min_row=2):
            ip_cell = row[ip_column_index - 1]
            if ip_cell.value and ip_cell.value.startswith(partial_ip):
                department_cell = row[department_column_index - 1]
                return department_cell.value

    # If the IP address is not found, return None
    return None

def find_department_email_recipients(department_name, department_file_path):
    # Load the department file
    department_workbook = openpyxl.load_workbook(department_file_path)
    department_sheet = department_workbook.active

    recipients = []

    # Find the department name in the department file
    for row in department_sheet.iter_rows(values_only=True):
        if department_name in row:
            department_column_index = row.index(department_name)

            # Retrieve all the row values of the department column
            for row_value in department_sheet.iter_rows(min_row=2, min_col=department_column_index + 1,
                                                        max_col=department_column_index + 1,
                                                        values_only=True):
                for recipient in row_value:
                    if recipient is not None:
                        recipients.append(recipient)
                    else:
                        break
            break

    return recipients


if len(sys.argv) < 1:
    print("Usage: python mail.py <subject> <body> <recipient1> <recipient2> ... <cc> <account_name>")
    sys.exit(1)

ip_file_path = "eGov.xlsx"
department_file_path = "Department_list.xlsx"
ip_address = sys.argv[1]
department_name = get_department_name(ip_file_path, ip_address)

recipients = find_department_email_recipients(department_name,department_file_path)
recipients = ";".join(recipients)

print(f"department name is {department_name}")

print( f'recipients of this department is {recipients}')